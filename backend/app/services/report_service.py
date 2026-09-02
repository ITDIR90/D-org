from datetime import datetime, timedelta, timezone

from collections import defaultdict

from io import BytesIO

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.enums import TaskStatus, UserRole
from app.core.permissions import get_user_admin_group_ids, is_admin_user
from app.models.task import Task
from app.models.user import User
from app.schemas.report import (
    CompletedTaskMatrixRow,
    CompletedTaskReportRow,
    CompletedTasksMatrixCell,
    CompletedTasksReport,
    EmployeeEfficiencyReport,
    EmployeeEfficiencyRow,
)


def _ensure_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _hours_between(start: datetime, end: datetime) -> float:
    return max(0.0, (_ensure_utc(end) - _ensure_utc(start)).total_seconds() / 3600)


async def get_employee_efficiency_report(
    db: AsyncSession,
    user: User,
    *,
    period_days: int = 30,
    group_id: int | None = None,
) -> EmployeeEfficiencyReport:
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Отчёт доступен только администраторам")
    if period_days < 1 or period_days > 365:
        raise HTTPException(status_code=400, detail="Период должен быть от 1 до 365 дней")

    since = datetime.now(timezone.utc) - timedelta(days=period_days)
    q = (
        select(Task)
        .options(selectinload(Task.assignee))
        .where(
            Task.status == TaskStatus.DONE,
            Task.completed_at.is_not(None),
            Task.completed_at >= since,
            Task.assignee_id.is_not(None),
        )
    )

    if group_id is not None:
        if user.role == UserRole.GROUP_ADMIN:
            admin_groups = await get_user_admin_group_ids(db, user)
            if group_id not in admin_groups:
                raise HTTPException(status_code=403, detail="Нет доступа к этой группе")
        q = q.where(Task.target_group_id == group_id)
    elif user.role == UserRole.GROUP_ADMIN:
        admin_groups = await get_user_admin_group_ids(db, user)
        if not admin_groups:
            return EmployeeEfficiencyReport(period_days=period_days, group_id=group_id, rows=[])
        q = q.where(Task.target_group_id.in_(admin_groups))

    result = await db.execute(q.order_by(Task.completed_at.desc()))
    tasks = result.scalars().all()

    stats: dict[int, dict] = defaultdict(
        lambda: {
            "full_name": "",
            "completed_count": 0,
            "completion_hours": [],
            "on_time_count": 0,
            "overdue_hours": [],
        }
    )

    for task in tasks:
        if not task.assignee_id or not task.completed_at:
            continue
        bucket = stats[task.assignee_id]
        bucket["full_name"] = task.assignee.full_name if task.assignee else f"#{task.assignee_id}"
        bucket["completed_count"] += 1
        hours = _hours_between(task.created_at, task.completed_at)
        bucket["completion_hours"].append(hours)
        due_at = _ensure_utc(task.due_at)
        completed_at = _ensure_utc(task.completed_at)
        if completed_at <= due_at:
            bucket["on_time_count"] += 1
        else:
            bucket["overdue_hours"].append(_hours_between(task.due_at, task.completed_at))

    rows: list[EmployeeEfficiencyRow] = []
    for user_id, data in stats.items():
        count = data["completed_count"]
        completion_hours = data["completion_hours"]
        overdue_hours = data["overdue_hours"]
        avg_completion = sum(completion_hours) / len(completion_hours) if completion_hours else None
        avg_overdue = sum(overdue_hours) / len(overdue_hours) if overdue_hours else None
        on_time_percent = round(data["on_time_count"] / count * 100, 1) if count else None
        rows.append(
            EmployeeEfficiencyRow(
                user_id=user_id,
                full_name=data["full_name"],
                completed_count=count,
                avg_completion_hours=round(avg_completion, 2) if avg_completion is not None else None,
                on_time_count=data["on_time_count"],
                on_time_percent=on_time_percent,
                avg_overdue_hours=round(avg_overdue, 2) if avg_overdue is not None else None,
            )
        )

    rows.sort(key=lambda r: (-r.completed_count, r.full_name.lower()))
    return EmployeeEfficiencyReport(period_days=period_days, group_id=group_id, rows=rows)


def _iso(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.astimezone(timezone.utc).isoformat()


async def _collect_completed_tasks(
    db: AsyncSession,
    user: User,
    *,
    date_from: datetime,
    date_to: datetime,
    group_id: int | None = None,
    user_id: int | None = None,
    category_id: int | None = None,
) -> list[Task]:
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Отчёт доступен только администраторам")

    q = (
        select(Task)
        .options(
            selectinload(Task.assignee),
            selectinload(Task.author),
            selectinload(Task.category),
            selectinload(Task.target_group),
        )
        .where(
            Task.status == TaskStatus.DONE,
            Task.completed_at.is_not(None),
            Task.completed_at >= date_from,
            Task.completed_at < date_to,
        )
    )

    if group_id is not None:
        if user.role == UserRole.GROUP_ADMIN:
            admin_groups = await get_user_admin_group_ids(db, user)
            if group_id not in admin_groups:
                raise HTTPException(status_code=403, detail="Нет доступа к этой группе")
        q = q.where(Task.target_group_id == group_id)
    elif user.role == UserRole.GROUP_ADMIN:
        admin_groups = await get_user_admin_group_ids(db, user)
        if not admin_groups:
            return []
        q = q.where(Task.target_group_id.in_(admin_groups))

    if user_id is not None:
        q = q.where(Task.assignee_id == user_id)

    if category_id is not None:
        q = q.where(Task.category_id == category_id)

    result = await db.execute(q.order_by(Task.completed_at.asc()))
    return list(result.scalars().all())


async def get_completed_tasks_report(
    db: AsyncSession,
    user: User,
    *,
    date_from: datetime,
    date_to: datetime,
    group_id: int | None = None,
    user_id: int | None = None,
    category_id: int | None = None,
    group_by: str = "category",
) -> CompletedTasksReport:
    if group_by not in ("category", "user"):
        raise HTTPException(status_code=400, detail="group_by должен быть 'category' или 'user'")

    tasks = await _collect_completed_tasks(
        db,
        user,
        date_from=date_from,
        date_to=date_to,
        group_id=group_id,
        user_id=user_id,
        category_id=category_id,
    )

    details: list[CompletedTaskReportRow] = []
    for task in tasks:
        completed = _ensure_utc(task.completed_at)
        details.append(
            CompletedTaskReportRow(
                task_id=task.id,
                number=task.number,
                title=task.title,
                completed_at=_iso(task.completed_at) or "",
                completed_day=completed.date().isoformat(),
                assignee_id=task.assignee_id,
                assignee_name=task.assignee.full_name if task.assignee else "—",
                author_id=task.author_id,
                author_name=task.author.full_name if task.author else "—",
                category_id=task.category_id,
                category_name=task.category.name if task.category else "—",
                group_id=task.target_group_id,
                group_name=task.target_group.name if task.target_group else "—",
                due_at=_iso(task.due_at),
                priority=task.priority.value if task.priority else None,
            )
        )

    # Все дни периода как колонки
    delta = (date_to.date() - date_from.date()).days
    days = [ (date_from.date() + timedelta(days=i)).isoformat() for i in range(delta + 1) ]
    day_index = {d: i for i, d in enumerate(days)}

    # Выбор строк матрицы: категории или пользователи (ответственные)
    if group_by == "user":
        row_names: dict[int, str] = {
            (task.assignee_id if task.assignee_id is not None else 0): (
                task.assignee.full_name if task.assignee else "Без ответственного"
            )
            for task in tasks
        }
        row_key = lambda t: t.assignee_id if t.assignee_id is not None else 0
    else:
        row_names = {
            task.category_id: (task.category.name if task.category else "Без категории")
            for task in tasks
        }
        row_key = lambda t: t.category_id

    # matrix[row_id][day] -> список task_id
    matrix: dict[int, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    for task in tasks:
        completed = _ensure_utc(task.completed_at)
        day = completed.date().isoformat()
        matrix[row_key(task)][day].append(task.id)

    day_totals = [0] * len(days)
    rows: list[CompletedTaskMatrixRow] = []
    for row_id in sorted(row_names, key=lambda r: row_names[r].lower()):
        cells: list[CompletedTasksMatrixCell] = []
        total = 0
        for day in days:
            ids = matrix[row_id].get(day, [])
            count = len(ids)
            day_totals[day_index[day]] += count
            total += count
            cells.append(CompletedTasksMatrixCell(day=day, count=count, task_ids=ids))
        rows.append(CompletedTaskMatrixRow(row_id=row_id, row_name=row_names[row_id], total=total, cells=cells))

    return CompletedTasksReport(
        date_from=date_from.date().isoformat(),
        date_to=date_to.date().isoformat(),
        group_id=group_id,
        user_id=user_id,
        category_id=category_id,
        group_by=group_by,
        total=len(details),
        days=days,
        rows=rows,
        day_totals=day_totals,
        tasks=details,
    )


_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_DAY_FILL = PatternFill("solid", fgColor="F2F2F2")
_TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
_HEADER_FONT = Font(bold=True)


def _format_datetime(iso_str: str | None) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.astimezone().strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return iso_str


def _format_date(iso_str: str) -> str:
    try:
        d = datetime.fromisoformat(iso_str)
        return d.strftime("%d.%m.%Y")
    except ValueError:
        return iso_str


def _write_matrix_sheet(wb: Workbook, report: CompletedTasksReport) -> None:
    ws = wb.active
    by = report.group_by or "category"
    ws.title = f"Матрица ({'сотрудники' if by == 'user' else 'категории'} x дни)"

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Отчёт по выполненным задачам"
    c.font = _TITLE_FONT
    c.fill = _TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    ws["A2"].value = (
        f"Период: {_format_date(report.date_from)} — {_format_date(report.date_to)}"
        f"    |    Выполнено задач: {report.total}"
    )
    ws["A2"].font = Font(italic=True)

    header_row = 3
    ws.cell(row=header_row, column=1, value="Категория / День" if by == "category" else "Сотрудник / День")
    for idx, day in enumerate(report.days, start=2):
        day_cell = ws.cell(row=header_row, column=idx, value=_format_date(day))
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_header = ws.cell(row=header_row, column=len(report.days) + 2, value="Итого")
    total_header.font = _HEADER_FONT

    for col_idx in range(1, len(report.days) + 3):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    ws.freeze_panes = "B4"

    row_idx = header_row
    for cat in report.rows:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=cat.row_name)
        for idx, cell_data in enumerate(cat.cells, start=2):
            cell = ws.cell(row=row_idx, column=idx, value=cell_data.count if cell_data.count else None)
            cell.alignment = Alignment(horizontal="center")
        total_cell = ws.cell(row=row_idx, column=len(report.days) + 2, value=cat.total)
        total_cell.font = _HEADER_FONT

    # Строка «Итого по дням»
    row_idx += 1
    total_row = ws.cell(row=row_idx, column=1, value="Итого по дням")
    total_row.font = _HEADER_FONT
    for idx, total in enumerate(report.day_totals, start=2):
        cell = ws.cell(row=row_idx, column=idx, value=total if total else None)
        cell.font = _HEADER_FONT
        cell.fill = _TOTAL_FILL
        cell.alignment = Alignment(horizontal="center")
    grand = ws.cell(row=row_idx, column=len(report.days) + 2, value=report.total)
    grand.font = _HEADER_FONT
    grand.fill = _TOTAL_FILL

    ws.column_dimensions["A"].width = 32
    for idx in range(2, len(report.days) + 3):
        ws.column_dimensions[get_column_letter(idx)].width = 12


def _write_detail_sheet(wb: Workbook, report: CompletedTasksReport) -> None:
    ws = wb.create_sheet("Список задач")

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Детализация: выполненные задачи"
    c.font = _TITLE_FONT
    c.fill = _TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = [
        "№ задачи",
        "Задача",
        "Дата выполнения",
        "Сотрудник",
        "Автор",
        "Категория",
        "Группа",
        "Приоритет",
    ]
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A4"

    row_idx = header_row
    for detail in report.tasks:
        row_idx += 1
        values = [
            detail.number,
            detail.title,
            _format_datetime(detail.completed_at),
            detail.assignee_name,
            detail.author_name,
            detail.category_name,
            detail.group_name,
            detail.priority or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = [10, 45, 18, 22, 22, 24, 22, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.auto_filter.ref = f"A{header_row}:H{row_idx}"


def export_completed_tasks_xlsx(report: CompletedTasksReport) -> tuple[bytes, str]:
    wb = Workbook()
    _write_matrix_sheet(wb, report)
    _write_detail_sheet(wb, report)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    by = "sotrudniki" if (report.group_by or "category") == "user" else "kategorii"
    filename = f"otchet-vypolnennye-zadachi-{by}-{report.date_from}-{report.date_to}.xlsx"
    return buffer.read(), filename
